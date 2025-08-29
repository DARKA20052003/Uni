import tkinter as tk
from openpyxl import Workbook
from openpyxl import load_workbook
import os

class Estudiante:
  def __init__(self, nombre, carrera, promedio):
    self.nombre=nombre
    self.carrera=carrera
    self.promedio=float(promedio)
class RegistroEstudiante:
  def __init__(self, archivo='/~/Uni/practicaPOO/aporte2/inventario.xlsx'):
    self.archivo=archivo
    self.estudiantes=[] #Inicia el set vacío de estudiantes como objeto
    self.cargarEstudiantes()

  def cargarEstudiantes(self):
    self.estudiantes=[]
    if os.path.exists(self.archivo):
      wb=load_workbook(self.archivo)
      ws=wb.active
      for row in ws.iter_rows(min_rows=2, values_only=True):
        if row[0] is not None:
          estudiante=Estudiante(row[0], row[1], row[2])
          self.estudiantes.append(estudiante)


