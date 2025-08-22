'''Registro de estudiantes
1- Definir una clase Estudiante con atributos: nombre, carrera y promedio.
2- Almacenar Estudiantes en un archivo Excel (estudiantes.xlsx) usando openpyxl.
3- Interfaz gráfica con tkinter para:
  - Ingresar estudiantes (nombre, carrera, promedio).
  - Guardarlos en el archivo Excel.
  - Mostrar la lista de estudiantes registrados en pantalla.
4- Incluir un botón que calcule y muestre el estudiante con el mayor promedio.
'''
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

class Estudiante:
    def __init__(self, nombre, carrera, promedio):
        self.nombre = nombre
        self.carrera = carrera
        self.promedio = float(promedio)  # Asegura tipo float

    def __str__(self):
        return f"{self.nombre} - {self.carrera} - {self.promedio}"

class RegistroEstudiantes:
    def __init__(self, archivo='estudiantes.xlsx'):
        self.archivo = archivo
        self.estudiantes = []
        self.cargar_estudiantes()

    def cargar_estudiantes(self):
        self.estudiantes = []  # Limpia la lista antes de cargar
        if os.path.exists(self.archivo):
            wb = load_workbook(self.archivo)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    estudiante = Estudiante(row[0], row[1], row[2])
                    self.estudiantes.append(estudiante)

    def agregar_estudiante(self, estudiante):
        self.estudiantes.append(estudiante)
        self.guardar_estudiantes()

    def guardar_estudiantes(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Carrera', 'Promedio'])
        for est in self.estudiantes:
            ws.append([est.nombre, est.carrera, est.promedio])
        wb.save(self.archivo)

    def estudiante_mejor_promedio(self):
        if not self.estudiantes:
            return None
        return max(self.estudiantes, key=lambda e: e.promedio)

# --- Interfaz gráfica ---
def actualizar_lista():
    lista.delete(0, tk.END)
    for est in registro.estudiantes:
        lista.insert(tk.END, str(est))

def agregar():
    nombre = entry_nombre.get()
    carrera = entry_carrera.get()
    promedio = entry_promedio.get()
    if not nombre or not carrera or not promedio:
        messagebox.showwarning("Campos vacíos", "Completa todos los campos.")
        return
    try:
        promedio_float = float(promedio)
    except ValueError:
        messagebox.showerror("Error", "El promedio debe ser un número.")
        return
    est = Estudiante(nombre, carrera, promedio_float)
    registro.agregar_estudiante(est)
    actualizar_lista()
    entry_nombre.delete(0, tk.END)
    entry_carrera.delete(0, tk.END)
    entry_promedio.delete(0, tk.END)

def mostrar_mejor():
    mejor = registro.estudiante_mejor_promedio()
    if mejor:
        messagebox.showinfo("Mejor promedio", f"{mejor.nombre} ({mejor.carrera}) - {mejor.promedio}")
    else:
        messagebox.showinfo("Mejor promedio", "No hay estudiantes registrados.")

registro = RegistroEstudiantes()

root = tk.Tk()
root.title("Registro de Estudiantes")

tk.Label(root, text="Nombre:").grid(row=0, column=0)
entry_nombre = tk.Entry(root)
entry_nombre.grid(row=0, column=1)

tk.Label(root, text="Carrera:").grid(row=1, column=0)
entry_carrera = tk.Entry(root)
entry_carrera.grid(row=1, column=1)

tk.Label(root, text="Promedio:").grid(row=2, column=0)
entry_promedio = tk.Entry(root)
entry_promedio.grid(row=2, column=1)

tk.Button(root, text="Agregar", command=agregar).grid(row=3, column=0, columnspan=2, pady=5)
tk.Button(root, text="Mejor Promedio", command=mostrar_mejor).grid(row=4, column=0, columnspan=2, pady=5)

lista = tk.Listbox(root, width=40)
lista.grid(row=5, column=0, columnspan=2, pady=10)
actualizar_lista()

root.mainloop()