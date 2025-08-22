'''Inventario de productos
1- Definir una clase producto con nombre, precio
y cantidad.
2- Guardar productos en un archivo Excel
(inventario.xlsx) usando openpyxl.
3- Interfaz gráfica con tkinter para:
  - Registrar productos (nombre, precio, cantidad).
  - Guardarlos en el arhcivo Excel.
  - Mostrar la lista de productos en pantalla.
4- Incluir un botón que calcule y muestre el valor total
del inventario (precio * cantidad).'''

import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os

# Clase Producto
class Producto:
    def __init__(self, nombre, precio, cantidad):
        self.nombre = nombre
        self.precio = float(precio)
        self.cantidad = int(cantidad)

    def valor_total(self):
        return self.precio * self.cantidad

# Funciones para manejar el archivo Excel
EXCEL_FILE = "/home/darka/Uni/practicaPOO/tallerFinal/Ejercicio2/inventario.xlsx"

def crear_archivo_si_no_existe():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre", "Precio", "Cantidad"])
        wb.save(EXCEL_FILE)

def guardar_producto(producto):
    crear_archivo_si_no_existe()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([producto.nombre, producto.precio, producto.cantidad])
    wb.save(EXCEL_FILE)

def cargar_productos():
    crear_archivo_si_no_existe()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    productos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        productos.append(Producto(*row))
    return productos

# Interfaz gráfica
class InventarioApp:
    def __init__(self, root):
        self.root = root
        root.title("Inventario de Productos")

        # Entradas
        tk.Label(root, text="Nombre:").grid(row=0, column=0)
        self.nombre_entry = tk.Entry(root)
        self.nombre_entry.grid(row=0, column=1)

        tk.Label(root, text="Precio:").grid(row=1, column=0)
        self.precio_entry = tk.Entry(root)
        self.precio_entry.grid(row=1, column=1)

        tk.Label(root, text="Cantidad:").grid(row=2, column=0)
        self.cantidad_entry = tk.Entry(root)
        self.cantidad_entry.grid(row=2, column=1)

        tk.Button(root, text="Registrar Producto", command=self.registrar_producto).grid(row=3, column=0, columnspan=2, pady=5)
        tk.Button(root, text="Mostrar Inventario", command=self.mostrar_inventario).grid(row=4, column=0, columnspan=2, pady=5)
        tk.Button(root, text="Calcular Valor Total", command=self.calcular_valor_total).grid(row=5, column=0, columnspan=2, pady=5)

        # Tabla de productos
        self.tree = ttk.Treeview(root, columns=("Nombre", "Precio", "Cantidad"), show="headings")
        self.tree.heading("Nombre", text="Nombre")
        self.tree.heading("Precio", text="Precio")
        self.tree.heading("Cantidad", text="Cantidad")
        self.tree.grid(row=6, column=0, columnspan=2, pady=10)

        # Valor total
        self.valor_total_label = tk.Label(root, text="Valor total: $0.00")
        self.valor_total_label.grid(row=7, column=0, columnspan=2)

    def registrar_producto(self):
        nombre = self.nombre_entry.get()
        precio = self.precio_entry.get()
        cantidad = self.cantidad_entry.get()
        if not nombre or not precio or not cantidad:
            messagebox.showwarning("Campos vacíos", "Por favor, completa todos los campos.")
            return
        try:
            producto = Producto(nombre, float(precio), int(cantidad))
            guardar_producto(producto)
            messagebox.showinfo("Éxito", "Producto registrado correctamente.")
            self.nombre_entry.delete(0, tk.END)
            self.precio_entry.delete(0, tk.END)
            self.cantidad_entry.delete(0, tk.END)
            self.mostrar_inventario()
        except ValueError:
            messagebox.showerror("Error", "Precio o cantidad inválidos.")

    def mostrar_inventario(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        productos = cargar_productos()
        for p in productos:
            self.tree.insert("", tk.END, values=(p.nombre, f"{p.precio:.2f}", p.cantidad))

    def calcular_valor_total(self):
        productos = cargar_productos()
        total = sum(p.valor_total() for p in productos)
        self.valor_total_label.config(text=f"Valor total: ${total:.2f}")

if __name__ == "__main__":
    crear_archivo_si_no_existe()
    root = tk.Tk()
    app = InventarioApp(root)
    root.mainloop()