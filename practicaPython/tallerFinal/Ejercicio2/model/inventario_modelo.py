'''Inventario de productos
1- Definir una clase producto con nombre, precio
y cantidad.
2- Guardar productos en un archivo Excel
(inventario.xlsx) usando openpyxl.
3- Interfaz gráfica con tkinter para:
  - Registrar productos (nombre, precio, cantidad).
  - Guardarlos en el arhcivo Excel.
  - Mostrar la lista de productos en pantalla.
4- Incluir un botón que calcule y muestre el valor total
del inventario (precio * cantidad).'''

from openpyxl import Workbook, load_workbook
import os

# Clase Producto
class Producto:
    def __init__(self, nombre, precio, cantidad):
        self.nombre = nombre
        self.precio = float(precio)
        self.cantidad = int(cantidad)

    def valor_total(self):
        return self.precio * self.cantidad

# Funciones para manejar el archivo Excel
EXCEL_FILE = "/home/darka/Uni/practicaPython/tallerFinal/Ejercicio2/repo/inventario.xlsx"

def crear_archivo_si_no_existe():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre", "Precio", "Cantidad"])
        wb.save(EXCEL_FILE)

def guardar_producto(producto):
    crear_archivo_si_no_existe()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([producto.nombre, producto.precio, producto.cantidad])
    wb.save(EXCEL_FILE)

def cargar_productos():
    crear_archivo_si_no_existe()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    productos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        productos.append(Producto(*row))
    return productos

