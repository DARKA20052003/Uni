'''Registro de estudiantes
1- Definir una clase Estudiante con atributos: nombre, carrera y promedio.
2- Almacenar Estudiantes en un archivo Excel (estudiantes.xlsx) usando openpyxl.
3- Interfaz gráfica con tkinter para:
  - Ingresar estudiantes (nombre, carrera, promedio).
  - Guardarlos en el archivo Excel.
  - Mostrar la lista de estudiantes registrados en pantalla.
4- Incluir un botón que calcule y muestre el estudiante con el mayor promedio.
'''
from openpyxl import Workbook, load_workbook
import os

class Estudiante:
    def __init__(self, nombre, carrera, promedio): #Clase Estudiante iniciada con nombre, carrera y promedio.
        self.nombre = nombre
        self.carrera = carrera
        self.promedio = float(promedio)  #Asegura que el promedio sea un número.

    def __str__(self):
        return f"{self.nombre} - {self.carrera} - {self.promedio}" #Retorna una cadena de la info del estudiante

class RegistroEstudiantes:
    def __init__(self, archivo='/home/darka/Uni/practicaPython/tallerFinal/Ejercicio1/repo/estudiantes.xlsx'): #Inicia el archivo en una clase con los métodos para cargar, agregar y calcular el mejor promedio.
        self.archivo = archivo
        self.estudiantes = [] #Set vacío de estudiantes.
        self.cargar_estudiantes()

    def cargar_estudiantes(self): #Carga la lista de estudiantes.
        self.estudiantes = []  # Limpia la lista antes de cargar
        if os.path.exists(self.archivo):
            wb = load_workbook(self.archivo)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    estudiante = Estudiante(row[0], row[1], row[2])
                    self.estudiantes.append(estudiante)

    def agregar_estudiante(self, estudiante):
        self.estudiantes.append(estudiante)
        if os.path.exists(self.archivo):
            wb = load_workbook(self.archivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Nombre', 'Carrera', 'Promedio'])
        ws.append([estudiante.nombre, estudiante.carrera, estudiante.promedio])
        wb.save(self.archivo)

    def estudiante_mejor_promedio(self):
        if not self.estudiantes:
            return None
        return max(self.estudiantes, key=lambda e: e.promedio)
